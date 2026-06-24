#!/usr/bin/env python3
"""
build_supply_chart.py — Automate the 5-mile new-construction Supply Chart.

Reads the three market exports (CoStar property roster, CoStar Data Analytics
time series, RealPage roster) for a subject property's 5-mile radius, reconciles
them into a single competitive-supply roster, buckets each property by lifecycle
stage, pins delivery quarters against the CoStar quarterly deliveries series, and
writes a formatted workbook with:
  - "Competitive Analysis": the colour-coded, chronological roster.
  - "Supply & Absorption": a relative-year (TTM) forecast of new supply,
    absorption, and overall occupancy with editable demand scenarios and pipeline
    toggles; plus subject rent/occupancy/concession rows (with --intake and
    --costar-subject-rents).
  - "Reconciliation Log": per-property merged values, sources, and conflicts.

See SKILL.md for the full methodology. Proximity (miles) is computed by geocoding
the subject and each comp (Nominatim, cached; offline ZIP-centroid fallback).
Verified lease-up rent/occupancy from HelloData is left flagged for the analyst.

Usage:
    python build_supply_chart.py \
        --subject-name "Canyon Ridge" \
        --subject-address "2552 E Gowen Rd" \
        --costar-roster   examples/canyon_ridge/CoStar_5mi_50unit_properties.xlsx \
        --costar-analytics examples/canyon_ridge/CoStar_5mi_Data_Analytics.xlsx \
        --realpage        examples/canyon_ridge/Realpage_5mi.xlsx \
        --out             output/Canyon_Ridge__Supply_Chart.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import geo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------- #
# Tunable thresholds (documented in SKILL.md)
# --------------------------------------------------------------------------- #
DEFAULT_STABILIZATION_TARGET = 0.95
# A delivered property is "stabilized" once occupancy >= this; recently delivered
# (within ~2 years) and below it = still "leasing up".
STABILIZED_OCC = 0.90
# How far back a delivery counts as "new construction" worth charting.
NEW_CONSTRUCTION_LOOKBACK_YEARS = 4

# --------------------------------------------------------------------------- #
# Styling (sampled from the reference template)
# --------------------------------------------------------------------------- #
NAVY = "FF1F3864"
BLUE = "FF2E75B6"
GRAY = "FFF2F2F2"
WHITE = "FFFFFFFF"
YELLOW = "FFFFFF00"

THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def font(bold=False, size=9, color="FF000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]


def quarter_index(year: int, q: int) -> int:
    """Absolute quarter index for ordering / arithmetic."""
    return year * 4 + (q - 1)


def fmt_quarter(year: int, q: int) -> str:
    return f"Q{q} {year}"


@dataclass
class Prop:
    name: str
    address: str
    units: Optional[int] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zipcode: Optional[str] = None
    year_built: Optional[int] = None
    construction_begin: Optional[str] = None
    status_raw: str = ""               # e.g. Existing / Stabilized / Pre-Planned
    costar_status: str = ""            # Existing / Under Construction / Proposed
    rp_status: str = ""                # Stabilized / Lease-Up / Under Construction / Pre-Planned / Planned
    # Per-source occupancy / rent so the two can be compared (not just merged).
    costar_occ: Optional[float] = None
    costar_rent: Optional[float] = None     # CoStar asking rent / unit
    rp_occ: Optional[float] = None
    rp_rent: Optional[float] = None         # RealPage effective rent / unit
    rp_asking: Optional[float] = None       # RealPage asking rent / unit (if exported)
    # Resolved primary values (chosen by source priority) used on the chart.
    occupancy: Optional[float] = None
    eff_rent: Optional[float] = None
    asking_rent: Optional[float] = None     # market/asking rent shown on the roster
    owner: Optional[str] = None
    stories: Optional[int] = None
    style: Optional[str] = None        # RealPage property style
    sources: set = field(default_factory=set)
    # derived
    est_delivery: Optional[str] = None     # "Q2 2023" or None (undated pipeline)
    deliv_year: Optional[int] = None
    deliv_q: Optional[int] = None
    bucket: Optional[str] = None
    prop_type: Optional[str] = None
    proximity_mi: Optional[float] = None   # straight-line miles subject -> comp
    roster_row: Optional[int] = None     # row on the Competitive Analysis sheet
    notes: list = field(default_factory=list)

    def note(self, txt: str):
        if txt and txt not in self.notes:
            self.notes.append(txt)


# --------------------------------------------------------------------------- #
# Address normalization & matching
# --------------------------------------------------------------------------- #
_STREET_NOISE = re.compile(
    r"\b(apartments?|apts?|townhomes?|the|at|on|of)\b", re.I)


def addr_key(address: str) -> Optional[str]:
    """Build a match key from a street address: leading number + first word.

    Handles ranges ("2410-2490 W Canal St" -> 2410) and directionals.
    """
    if not address:
        return None
    a = str(address).strip().lower()
    m = re.match(r"(\d+)", a)
    if not m:
        return None
    number = m.group(1)
    rest = a[m.end():].strip(" -")
    # drop a trailing range number like "-2490"
    rest = re.sub(r"^\d+\s*", "", rest)
    toks = [t for t in re.split(r"[\s,]+", rest) if t]
    # skip leading directional (n/s/e/w)
    dirs = {"n", "s", "e", "w", "ne", "nw", "se", "sw",
            "north", "south", "east", "west"}
    sig = next((t for t in toks if t not in dirs), toks[0] if toks else "")
    return f"{number} {sig}"


def name_key(name: str) -> str:
    n = _STREET_NOISE.sub(" ", str(name or "").lower())
    return re.sub(r"\s+", " ", n).strip()


# --------------------------------------------------------------------------- #
# Parsers
# --------------------------------------------------------------------------- #
def _hdr_map(ws):
    headers = {}
    for j, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
        if c.value is not None:
            headers[str(c.value).strip()] = j
    return headers


def _int(v):
    try:
        if v in (None, "", "-", "—"):
            return None
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def _float(v):
    try:
        if v in (None, "", "-", "—"):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_costar_roster(path) -> list[Prop]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    h = _hdr_map(ws)

    def col(row, key):
        j = h.get(key)
        return row[j - 1] if j else None

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = col(row, "Property Name")
        if not name:
            continue
        # CoStar v2 adds asking rent and vacancy; v1 lacks them.
        rent = _float(col(row, "Avg Asking/Unit"))
        vac = _float(col(row, "Vacancy %"))   # whole-percent, e.g. 8 -> 8%
        occ = (1 - vac / 100.0) if vac is not None else None
        p = Prop(
            name=str(name).strip(),
            address=str(col(row, "Property Address") or "").strip(),
            city=(str(col(row, "City")).strip() if col(row, "City") else None),
            state=(str(col(row, "State")).strip() if col(row, "State") else None),
            zipcode=(str(col(row, "Zip")).strip()[:5] if col(row, "Zip") else None),
            units=_int(col(row, "Number of Units")),
            year_built=_int(col(row, "Year Built")),
            construction_begin=(str(col(row, "Construction Begin")).strip()
                                if col(row, "Construction Begin") else None),
            status_raw=str(col(row, "Building Status") or "").strip(),
            costar_status=str(col(row, "Building Status") or "").strip(),
            costar_occ=occ,
            costar_rent=rent,
            owner=(str(col(row, "Owner Name")).strip()
                   if col(row, "Owner Name") else None),
            stories=_int(col(row, "Number of Stories")),
        )
        p.sources.add("CoStar")
        out.append(p)
    return out


def parse_realpage(path) -> list[Prop]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    h = _hdr_map(ws)

    def col(row, key):
        j = h.get(key)
        return row[j - 1] if j else None

    def col_any(row, keys):
        for k in keys:
            v = col(row, k)
            if v is not None:
                return v
        return None

    # RealPage exports effective rent; some pulls also carry an asking/market
    # rent column. Capture it if present (never use effective rent as "asking").
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = col(row, "Name")
        if not name:
            continue
        p = Prop(
            name=str(name).strip(),
            address=str(col(row, "Address") or "").strip(),
            city=(str(col(row, "City")).strip() if col(row, "City") else None),
            state=(str(col(row, "State")).strip() if col(row, "State") else None),
            zipcode=(str(col(row, "Zip Code")).strip()[:5] if col(row, "Zip Code") else None),
            units=_int(col(row, "Total Units")),
            year_built=_int(col(row, "Year Built")),
            status_raw=str(col(row, "Property Status") or "").strip(),
            rp_status=str(col(row, "Property Status") or "").strip(),
            rp_occ=_float(col(row, "Occupancy")),
            rp_rent=_float(col(row, "Effective Rent")),
            rp_asking=_float(col_any(row, ("Asking Rent", "Market Rent",
                                           "Asking Rent / Unit", "Avg Asking Rent"))),
            owner=(str(col(row, "Property Owner")).strip()
                   if col(row, "Property Owner") else None),
            stories=_int(col(row, "Stories")),
            style=(str(col(row, "Property Style")).strip()
                   if col(row, "Property Style") else None),
        )
        p.sources.add("RealPage")
        out.append(p)
    return out


_QRE = re.compile(r"(\d{4})\s*Q([1-4])")


def parse_costar_analytics(path):
    """Parse the CoStar Data Analytics quarterly time series.

    Returns (latest_inventory_units, deliveries{(y,q):units}, latest_label,
    series{(y,q): {...}}) where each series entry has inventory, deliveries,
    occupancy_units, absorption, occupancy_pct for that quarter.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    h = _hdr_map(ws)
    pcol = h.get("Period")
    cols = {
        "inventory": h.get("Inventory Units"),
        "deliveries": h.get("Deliveries Units"),
        "occ_units": h.get("Occupancy Units"),
        "absorption": h.get("Absorption Units"),
        "occ_pct": h.get("Occupancy Percent"),
        "uc_units": h.get("Under Construction Units"),
        "eff_rent": h.get("Effective Rent Per Unit"),
    }
    deliveries = {}
    series = {}
    latest_inv = None
    latest_label = None
    latest_uc = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        period = row[pcol - 1] if pcol else None
        if not period:
            continue
        m = _QRE.search(str(period))
        if not m:
            continue
        year, q = int(m.group(1)), int(m.group(2))
        rec = {k: (_float(row[c - 1]) if c else None) for k, c in cols.items()}
        series[(year, q)] = rec
        d = _int(rec["deliveries"])
        if d:
            deliveries[(year, q)] = d
        if latest_inv is None:  # rows are newest-first
            latest_inv = _int(rec["inventory"])
            latest_label = str(period).strip()
            latest_uc = _int(rec["uc_units"])
    return latest_inv, deliveries, latest_label, series, latest_uc


# --------------------------------------------------------------------------- #
# Reconciliation
# --------------------------------------------------------------------------- #
def reconcile(costar: list[Prop], realpage: list[Prop]) -> list[Prop]:
    """Merge the two rosters by address (primary) / name (fallback)."""
    merged: dict[str, Prop] = {}
    by_name: dict[str, str] = {}

    def same_property(a: Prop, b: Prop) -> bool:
        """Guard against merging two distinct properties that share an address
        (e.g. an existing asset and its planned redevelopment on the same site)."""
        ta = {t for t in name_key(a.name).split() if len(t) >= 4}
        tb = {t for t in name_key(b.name).split() if len(t) >= 4}
        if ta & tb:
            return True                       # share a real name token -> same
        # no name overlap: an existing asset vs a pipeline deal = different
        a_pipe = _proposed_signal(a) or _uc_signal(a)
        b_pipe = _proposed_signal(b) or _uc_signal(b)
        if (a_pipe and _existing_signal(b)) or (b_pipe and _existing_signal(a)):
            return False
        if a.year_built and b.year_built and abs(a.year_built - b.year_built) > 5:
            return False
        if a.units and b.units and abs(a.units - b.units) > 0.25 * max(a.units, b.units):
            return False
        return True

    def register(p: Prop):
        ak = addr_key(p.address)
        if ak and ak in merged and same_property(merged[ak], p):
            return merge_into(merged[ak], p)
        nk = name_key(p.name)
        if nk in by_name:
            cand = merged.get(by_name[nk])
            if cand is not None and same_property(cand, p):
                return merge_into(cand, p)
        key = ak or ("name:" + nk)
        if key in merged:                     # address taken by a different property
            key = f"{key}|{nk}"
        merged[key] = p
        by_name.setdefault(nk, key)
        return p

    def merge_into(base: Prop, other: Prop):
        base.sources |= other.sources
        # Units: keep CoStar's; flag if they differ materially
        if other.units and base.units and other.units != base.units:
            lo, hi = sorted((base.units, other.units))
            if hi - lo >= 3:  # ignore +/-2 unit noise
                base.note(f"Units: CoStar {base.units if 'CoStar' in base.sources else hi} "
                          f"vs RealPage {other.units if 'RealPage' in other.sources else lo}")
        # Occupancy / rent: keep both sources' values for comparison.
        base.costar_occ = base.costar_occ if base.costar_occ is not None else other.costar_occ
        base.costar_rent = base.costar_rent if base.costar_rent is not None else other.costar_rent
        base.rp_occ = base.rp_occ if base.rp_occ is not None else other.rp_occ
        base.rp_rent = base.rp_rent if base.rp_rent is not None else other.rp_rent
        base.rp_asking = base.rp_asking if base.rp_asking is not None else other.rp_asking
        # Style / stories / owner / location backfill
        base.style = base.style or other.style
        base.stories = base.stories or other.stories
        base.owner = base.owner or other.owner
        base.city = base.city or other.city
        base.state = base.state or other.state
        base.zipcode = base.zipcode or other.zipcode
        # Year built conflict
        if other.year_built and base.year_built and other.year_built != base.year_built:
            base.note(f"Year built: {min(base.year_built, other.year_built)}/"
                      f"{max(base.year_built, other.year_built)}")
        base.year_built = base.year_built or other.year_built
        base.construction_begin = base.construction_begin or other.construction_begin
        # Carry both sources' lifecycle status (used by classify()).
        base.costar_status = base.costar_status or other.costar_status
        base.rp_status = base.rp_status or other.rp_status
        return base

    for p in costar:
        register(p)
    for p in realpage:
        register(p)
    return list(merged.values())


OCC_DIVERGENCE = 0.02   # >=2 pts occupancy gap between sources -> flag
RENT_DIVERGENCE = 0.05  # >=5% rent gap between sources -> flag


def resolve_occ_rent(p: Prop, occ_source: str, rent_source: str,
                     flag_divergence: bool = True):
    """Choose the displayed occupancy/rent by source priority; flag divergence.

    CoStar rent is *asking*; RealPage rent is *effective* — divergence on rent is
    expected and is surfaced rather than silently averaged. Divergence notes are
    only added for delivered assets (a partial-lease-up vs 0% gap on a
    not-yet-open building is noise, not signal).
    """
    prio = {"costar": (p.costar_occ, p.rp_occ), "realpage": (p.rp_occ, p.costar_occ)}
    p.occupancy = next((v for v in prio[occ_source] if v is not None), None)

    # Roster "Avg Mkt Rent" is an ASKING/market figure. CoStar asking is the
    # primary source; RealPage's only rent is *effective* (not asking), so it is
    # never used here — RealPage contributes only when its export carries an
    # asking column (p.rp_asking). 'average' blends the available asking values.
    co, rp = p.costar_rent, p.rp_asking
    if rent_source == "realpage":
        p.asking_rent = rp if rp is not None else co
    elif rent_source == "average":
        vals = [v for v in (co, rp) if v is not None]
        p.asking_rent = sum(vals) / len(vals) if vals else None
    else:  # costar (default)
        p.asking_rent = co if co is not None else rp
    p.eff_rent = p.rp_rent          # effective rent retained for the Recon log

    if not flag_divergence:
        return
    if p.costar_occ is not None and p.rp_occ is not None \
            and abs(p.costar_occ - p.rp_occ) >= OCC_DIVERGENCE:
        p.note(f"Occ: CoStar {p.costar_occ:.0%} vs RealPage {p.rp_occ:.0%}")
    if p.costar_rent and p.rp_rent \
            and abs(p.costar_rent - p.rp_rent) / max(p.costar_rent, p.rp_rent) >= RENT_DIVERGENCE:
        p.note(f"Rent: CoStar ask ${p.costar_rent:,.0f} vs "
               f"RealPage eff ${p.rp_rent:,.0f}")


def _existing_signal(p: Prop) -> bool:
    """Either source says the building physically exists / is delivered."""
    return (p.costar_status.lower() == "existing"
            or p.rp_status.lower() in ("stabilized", "lease-up", "lease up"))


def _uc_signal(p: Prop) -> bool:
    cs = p.costar_status.lower()
    rp = p.rp_status.lower()
    return ("under construction" in cs
            or rp in ("under construction", "under construction/lease-up"))


def _proposed_signal(p: Prop) -> bool:
    cs = p.costar_status.lower()
    rp = p.rp_status.lower()
    return cs == "proposed" or rp in ("pre-planned", "preplanned", "planned",
                                      "proposed")


def _is_pipeline(p: Prop) -> bool:
    """A forward-supply deal (under construction or proposed), not yet delivered."""
    return (_uc_signal(p) or _proposed_signal(p)) and not _existing_signal(p)


# --------------------------------------------------------------------------- #
# Delivery-quarter pinning + bucketing
# --------------------------------------------------------------------------- #
def pin_delivery(p: Prop, deliveries: dict):
    """Assign a delivery quarter to a *delivered* property.

    Match the property's unit count to the CoStar quarterly deliveries series
    within +/-1 year of its year-built. An exact unit match pins the quarter; with
    no exact match we keep the year only ("Q? <year>", quarter unknown) rather than
    guessing — guessed quarters are noise in a large market. Undated delivered
    deals can be quarter-stamped by the analyst via --pipeline-dates.
    """
    if not p.year_built or not p.units:
        if p.year_built:
            p.deliv_year = p.year_built
            p.est_delivery = f"Q? {p.year_built}"
        return
    # 1) Exact unit match within +/-1 year -> precise quarter.
    exact = [(y, q) for (y, q), u in deliveries.items()
             if abs(y - p.year_built) <= 1 and u == p.units]
    if exact:
        y, q = sorted(exact)[0]
        p.deliv_year, p.deliv_q = y, q
        p.est_delivery = fmt_quarter(y, q)
        return
    # 2) No exact match: pick the same-year quarter with the closest delivered
    #    count (estimated). Keeps the absorption table populated; flagged as est.
    same_year = [(q, u) for (y, q), u in deliveries.items()
                 if y == p.year_built and u > 0]
    if same_year:
        q, _ = min(same_year, key=lambda t: abs(t[1] - p.units))
        p.deliv_year, p.deliv_q = p.year_built, q
        p.est_delivery = fmt_quarter(p.year_built, q)
        p.note("Delivery quarter estimated")
        return
    # 3) Year known but no deliveries recorded that year -> year only.
    p.deliv_year = p.year_built
    p.est_delivery = f"Q? {p.year_built}"


_QLABEL = re.compile(r"Q([1-4])\s*'?(\d{2,4})|(\d{4})\s*Q([1-4])", re.I)


def parse_quarter_label(label: str) -> Optional[tuple[int, int]]:
    """Parse 'Q2 2028' or '2028 Q2' (or Q2'28) -> (year, quarter)."""
    if not label:
        return None
    m = _QLABEL.search(str(label).strip())
    if not m:
        return None
    if m.group(1):
        q = int(m.group(1)); y = int(m.group(2))
        if y < 100:
            y += 2000
    else:
        y = int(m.group(3)); q = int(m.group(4))
    return y, q


def load_pipeline_dates(path) -> dict:
    """Read analyst-supplied delivery dates for pipeline deals.

    CSV with headers: property, est_delivery[, units]. Returns
    {name_key: {"yq": (y,q), "units": int|None}}.
    """
    import csv
    out = {}
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            r = {(k or "").strip().lower(): (v or "").strip()
                 for k, v in row.items()}
            name = r.get("property")
            yq = parse_quarter_label(r.get("est_delivery", ""))
            if not name or not yq:
                continue
            out[name_key(name)] = {"yq": yq, "units": _int(r.get("units"))}
    return out


def apply_pipeline_dates(props: list[Prop], dates: dict):
    for p in props:
        info = dates.get(name_key(p.name))
        if not info:
            continue
        y, q = info["yq"]
        p.deliv_year, p.deliv_q = y, q
        p.est_delivery = fmt_quarter(y, q)
        if info["units"]:
            p.units = info["units"]
        p.note("Est. delivery set by analyst")


def emit_pipeline_template(props: list[Prop], path):
    """Write a CSV of undated pipeline deals for the analyst to fill in."""
    import csv
    # Forecast-relevant deals without a precise delivery quarter.
    undated = [p for p in props
               if p.deliv_q is None
               and p.bucket in ("LEASING UP", "UNDER CONSTRUCTION", "PROPOSED")]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["property", "est_delivery", "units", "bucket", "costar_year",
                    "# Fill est_delivery as 'Q2 2028'. Re-run with --pipeline-dates."])
        for p in sorted(undated, key=lambda x: (x.bucket, -(x.units or 0))):
            w.writerow([p.name, "", p.units or "", p.bucket, p.year_built or "", ""])
    return len(undated)


DILIGENCE_COLS = ["type", "property", "units", "est_delivery", "status",
                  "leasing_pace", "notes", "source"]


def emit_diligence_template(props: list[Prop], path):
    """Write a per-project research template (the SKILL.md research phase fills
    it). Lists every lease-up / UC / proposed deal to verify, plus a blank row to
    add shadow-supply sites (rezonings, entitled/vacant tracts, untracked deals)."""
    import csv
    targets = [p for p in props
               if p.bucket in ("LEASING UP", "UNDER CONSTRUCTION", "PROPOSED")]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(DILIGENCE_COLS)
        for p in sorted(targets, key=lambda x: (x.bucket, -(x.units or 0))):
            w.writerow(["pipeline", p.name, p.units or "", p.est_delivery or "", "",
                        "", f"(verify {p.bucket.lower()})", ""])
        w.writerow(["shadow", "<latent site / rezoning / land sale>", "", "", "",
                    "", "potential future supply not in CoStar/RealPage", ""])
    return len(targets)


def load_diligence(path):
    """Read a filled diligence CSV (DILIGENCE_COLS). Returns list of dict rows."""
    import csv
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            r = {(k or "").strip().lower(): (v or "").strip()
                 for k, v in row.items()}
            if r.get("property") and not r["property"].startswith("<"):
                rows.append(r)
    return rows


def apply_diligence(props: list[Prop], rows):
    """Fold researched delivery/unit corrections back into the pipeline props."""
    by = {name_key(p.name): p for p in props}
    for r in rows:
        if r.get("type", "pipeline") != "pipeline":
            continue
        p = by.get(name_key(r["property"]))
        if not p:
            continue
        yq = parse_quarter_label(r.get("est_delivery", ""))
        if yq:
            p.deliv_year, p.deliv_q = yq
            p.est_delivery = fmt_quarter(*yq)
        if _int(r.get("units")):
            p.units = _int(r["units"])
        st = (r.get("status") or "").lower()
        if st:
            p.note(f"Diligence: {r['status']}")
            # researched status corrects the auto-classification
            if any(k in st for k in ("under construction", "leasing", "delivered")) \
                    and p.bucket == "PROPOSED":
                p.bucket = "UNDER CONSTRUCTION"
            elif any(k in st for k in ("proposed", "permitted", "planned",
                                       "stall", "cancel", "entitle")) \
                    and p.bucket == "UNDER CONSTRUCTION":
                p.bucket = "PROPOSED"



def _find_row(ws, needle, col=1, maxr=60):
    needle = needle.lower()
    for r in range(1, maxr + 1):
        v = ws.cell(r, col).value
        if v and needle in str(v).lower():
            return r
    return None


def parse_intake_subject(path):
    """Subject monthly market rent / effective rent / occupancy from the RR-T12
    intake's 'Lease Trend' tab. Returns {(year, month): {mkt, eff, occ}}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Lease Trend" not in wb.sheetnames:
        return {}
    ws = wb["Lease Trend"]
    hdr_row = next((r for r in range(1, 30)
                    if str(ws.cell(r, 1).value or "").strip().lower().startswith("month")),
                   None)
    if not hdr_row:
        return {}
    mkt_row = _find_row(ws, "HD Market Rent")
    eff_row = _find_row(ws, "HD Effective Rent / unit") or _find_row(ws, "HD Effective Rent")
    occ_row = _find_row(ws, "Physical Occupancy")
    conc_row = _find_row(ws, "HD Concession %")
    out = {}
    for c in range(2, ws.max_column + 1):
        my = parse_month_year(ws.cell(hdr_row, c).value)
        if not my:
            continue
        rec = {}
        for key, rw in (("mkt", mkt_row), ("eff", eff_row), ("occ", occ_row),
                        ("conc", conc_row)):
            if rw:
                v = ws.cell(rw, c).value
                if isinstance(v, (int, float)):
                    rec[key] = v
        if rec:
            out[my] = rec
    return out


def parse_realpage_subject_rents(path, subject_name):
    """Subject quarterly asking/effective rent from a RealPage per-property
    10-yr export (wide format: Metric rows, Y####Q# columns). Returns
    {(year, q): {ask, eff}}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    qcols = {}
    for c, v in enumerate(hdr, 1):
        m = re.match(r"Y(\d{4})Q([1-4])", str(v or ""))
        if m:
            qcols[c] = (int(m.group(1)), int(m.group(2)))
    key = name_key(subject_name)
    out = {}
    for r in range(2, ws.max_row + 1):
        if name_key(str(ws.cell(r, 1).value or "")) != key:
            continue
        metric = str(ws.cell(r, 4).value or "").strip().lower()
        field = "eff" if "effective" in metric else "ask" if "asking" in metric else None
        if not field:
            continue
        for c, yq in qcols.items():
            v = _float(ws.cell(r, c).value)
            if v:
                out.setdefault(yq, {}).setdefault(field, v)
    return out


def parse_costar_subject_rents(path, subject_name):
    """Subject quarterly asking/effective rent from the CoStar per-property
    (50-unit) analytics export. Returns {(year, q): {ask, eff}}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    row1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    hdr = {(v or "").strip(): i for i, v in enumerate(row2, 1)}

    def grouped(group, sub):
        cur = None
        for i, (a, b) in enumerate(zip(row1, row2), 1):
            if a:
                cur = a
            if cur == group and b == sub:
                return i
        return None

    nmc = hdr.get("Building Name") or hdr.get("Property Name")
    pc = hdr.get("Period")
    ask_c = grouped("Asking Rent", "Per Unit")
    eff_c = grouped("Effective Rent", "Per Unit")
    occ_c = grouped("Occupancy", "Percent")
    conc_c = hdr.get("Concessions %")
    key = name_key(subject_name)
    out = {}
    for r in range(3, ws.max_row + 1):
        nm = ws.cell(r, nmc).value if nmc else None
        if not nm or name_key(str(nm)) != key:
            continue
        m = _QRE.search(str(ws.cell(r, pc).value)) if pc else None
        if not m:
            continue
        yq = (int(m.group(1)), int(m.group(2)))
        out[yq] = {"ask": _float(ws.cell(r, ask_c).value) if ask_c else None,
                   "eff": _float(ws.cell(r, eff_c).value) if eff_c else None,
                   "occ": _float(ws.cell(r, occ_c).value) if occ_c else None,
                   "conc": _float(ws.cell(r, conc_c).value) if conc_c else None}
    return out


_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct",
     "nov", "dec"], start=1)}


def parse_month_year(s):
    """Parse 'Nov 2023' / 'November 2023' -> (year, month) or None."""
    if not s:
        return None
    m = re.search(r"([A-Za-z]{3,})\s+(\d{4})", str(s))
    if not m:
        return None
    mon = _MONTHS.get(m.group(1)[:3].lower())
    if not mon:
        return None
    return int(m.group(2)), mon


CONSTRUCTION_MONTHS = 24   # typical build duration: begin -> delivery


def estimate_pipeline_delivery(p: Prop, as_of: tuple[int, int]):
    """Estimate a delivery quarter for an under-construction / proposed deal.

    Priority: (1) Construction Begin + ~24 months -> quarter; (2) CoStar Year
    Built (= expected completion year) at mid-year. Always pushed to land in the
    future (after the as-of quarter) so it enters the forward forecast.
    """
    est_idx = None
    my = parse_month_year(p.construction_begin)
    if my:
        y, mo = my
        mo += CONSTRUCTION_MONTHS
        y += (mo - 1) // 12
        mo = (mo - 1) % 12 + 1
        q = (mo - 1) // 3 + 1
        est_idx = quarter_index(y, q)
        # reconcile with CoStar completion year if it disagrees by >1 yr
        if p.year_built and abs(y - p.year_built) > 1:
            est_idx = quarter_index(p.year_built, q)
    elif p.year_built:
        est_idx = quarter_index(p.year_built, 2)
    if est_idx is None:
        return
    est_idx = max(est_idx, quarter_index(*as_of) + 1)
    p.deliv_year, p.deliv_q = est_idx // 4, est_idx % 4 + 1
    p.est_delivery = fmt_quarter(p.deliv_year, p.deliv_q)


def classify(p: Prop, as_of: tuple[int, int], target: float):
    """Assign one of the four lifecycle buckets, driven by source status.

    Precedence: a delivered/existing signal (from either source) wins over
    under-construction, which wins over proposed. Among delivered deals, occupancy
    and recency split stabilized vs leasing-up.
    """
    recent = bool(p.year_built and p.year_built >= as_of[0] - 2)
    # "Started" = construction has broken ground on/before the as-of quarter.
    my = parse_month_year(p.construction_begin)
    begin_idx = quarter_index(my[0], (my[1] - 1) // 3 + 1) if my else None
    started = begin_idx is not None and begin_idx <= quarter_index(*as_of)

    if _existing_signal(p):
        occ = p.occupancy
        if occ is None:
            p.bucket = "LEASING UP" if recent else "STABILIZED / STABILIZING"
        elif occ >= STABILIZED_OCC:
            p.bucket = "STABILIZED / STABILIZING"
        elif recent:
            p.bucket = "LEASING UP"
        else:
            p.bucket = "STABILIZED / STABILIZING"   # older underperformer
    elif _uc_signal(p) or started:
        # Under construction = a UC status from either source, OR ground already
        # broken (construction begin on/before as-of) even if a source lags.
        p.bucket = "UNDER CONSTRUCTION"
    elif _proposed_signal(p):
        p.bucket = "PROPOSED"
    elif p.deliv_q is not None:                      # exact delivery, no status
        p.bucket = "LEASING UP" if recent and (p.occupancy or 0) < STABILIZED_OCC \
            else "STABILIZED / STABILIZING"
    elif p.year_built and p.year_built > as_of[0]:
        p.bucket = "UNDER CONSTRUCTION"
    elif recent:
        p.bucket = "LEASING UP"
    else:
        p.bucket = "STABILIZED / STABILIZING"

    # Lease-ups are where rent/occ accuracy matters most -> flag for HelloData
    if p.bucket == "LEASING UP":
        p.note("Verify rent/occ w/ HelloData")
    # Pipeline assets carry no real rent/occ yet
    if p.bucket in ("UNDER CONSTRUCTION", "PROPOSED"):
        p.occupancy = None
        p.eff_rent = None
    if p.bucket == "PROPOSED" and "RealPage" in p.sources and "CoStar" not in p.sources:
        p.note("RealPage pipeline (not yet in CoStar)")
    if p.units and p.units < 50 and "CoStar" not in p.sources:
        p.note("RealPage only (sub-50 unit)")


def derive_type(p: Prop) -> str:
    style = (p.style or "").lower()
    if "town" in style or "town" in p.name.lower():
        return "Townhome"
    if "single" in style:
        return "Single Family"
    if style in ("podium", "wrap"):
        return "Mid-Rise"
    if style == "garden":
        # refine by stories
        if p.stories and p.stories >= 4:
            return "Mid-Rise"
        return "Garden"
    s = p.stories or 0
    if s >= 8:
        return "High-Rise"
    if s >= 4:
        return "Mid-Rise"
    if s == 3:
        return "Low-Rise"
    return "Garden"


# --------------------------------------------------------------------------- #
# Workbook writer
# --------------------------------------------------------------------------- #
BUCKET_ORDER = [
    ("STABILIZED / STABILIZING", "Stabilized / leasing-complete deliveries"),
    ("LEASING UP", "Recently delivered, still leasing"),
    ("UNDER CONSTRUCTION", "Under construction, not yet delivered"),
    ("PROPOSED", "Proposed / pre-planned pipeline"),
]

# Per-bucket section colours, carried over from the reference template:
# header band colour + a light matching row tint to distinguish each section.
BUCKET_STYLES = {
    "STABILIZED / STABILIZING": {"header": "FF2E75B6", "row": "FFDDEBF7"},  # blue
    "LEASING UP":               {"header": "FFED7D31", "row": "FFFCE4D6"},  # orange
    "UNDER CONSTRUCTION":       {"header": "FF375623", "row": "FFE2EFDA"},  # green
    "PROPOSED":                 {"header": "FFFF0000", "row": "FFFCE4E4"},  # red
}

COLS = {  # column letter -> (header, width)
    "B": ("#", 3.5),
    "C": ("Property", 30),
    "D": ("Units", 8),
    "E": ("Est. Delivery", 11.5),
    "F": ("Occupancy", 10),
    "G": ("Avg Mkt Rent", 13),
    "H": ("Owner", 22),
    "I": ("Proximity", 11),
    "J": ("Type", 14),
    "K": ("Notes", 30),
}


def hist_annual_absorption(series: dict, as_of: tuple[int, int], n_qtrs=12):
    """Average annualized net absorption over the trailing n quarters (CoStar)."""
    as_idx = quarter_index(*as_of)
    vals = [rec["absorption"] for (y, q), rec in series.items()
            if 0 <= as_idx - quarter_index(y, q) < n_qtrs
            and rec.get("absorption") is not None]
    if not vals:
        return 0
    return sum(vals) * 4 / len(vals)


def subject_annual_by_window(plan, as_idx, subj_monthly, subj_costar,
                             subj_realpage=None):
    """Aggregate subject rents/occupancy into each relative-year TTM window.

    HelloData (mix-weighted) is preferred; where it doesn't cover a window fall
    back to CoStar or RealPage per-property rents — whichever tracks HelloData
    more closely in the overlap (chosen per deal/market) — level-aligned via the
    overlap ratio. Occupancy fallback always comes from CoStar (RealPage 10-yr has
    no occupancy). Returns {column_label: {mkt,eff,occ,conc,src}} and the chosen
    fallback source name.
    """
    out = {}
    subj_monthly = subj_monthly or {}
    subj_costar = subj_costar or {}
    subj_realpage = subj_realpage or {}

    def cal_year_avg(d, key, idxfn):
        agg = {}
        for k, rec in d.items():
            if rec.get(key) is not None:
                agg.setdefault(idxfn(k), []).append(rec[key])
        return {y: sum(v) / len(v) for y, v in agg.items()}
    hd_eff_y = cal_year_avg(subj_monthly, "eff", lambda k: k[0])

    def fit(src):
        sy = cal_year_avg(src, "eff", lambda k: k[0])
        common = [y for y in hd_eff_y if y in sy and sy[y]]
        if not common:
            return (float("inf"), 1.0)
        miss = sum(abs(hd_eff_y[y] - sy[y]) for y in common) / len(common)
        ratio = sum(hd_eff_y[y] / sy[y] for y in common) / len(common)
        return (miss, ratio)
    cs_fit, rp_fit = fit(subj_costar), fit(subj_realpage)
    # pick the better-tracking source for rents (CoStar wins ties / when no HD)
    if rp_fit[0] < cs_fit[0]:
        fb, ratio, fb_name = subj_realpage, rp_fit[1], "RealPage"
    else:
        fb, ratio, fb_name = subj_costar, cs_fit[1], "CoStar"

    for (label, kind, k) in plan:
        if kind != "hist":
            continue
        end_idx = as_idx - 4 * k
        ey, eq = end_idx // 4, end_idx % 4 + 1
        end_m = eq * 3
        months = []
        y, m = ey, end_m
        for _ in range(12):
            months.append((y, m))
            m -= 1
            if m == 0:
                m = 12; y -= 1
        hd = [subj_monthly[mm] for mm in months if mm in subj_monthly]
        def avg(recs, key):
            vals = [r[key] for r in recs if r.get(key) is not None]
            return sum(vals) / len(vals) if vals else None
        qs = [(yy, qq) for yy in (ey, ey - 1) for qq in (1, 2, 3, 4)
              if end_idx - 4 < quarter_index(yy, qq) <= end_idx]
        def srcavg(source, key):
            vals = [source[q][key] for q in qs
                    if q in source and source[q].get(key) is not None]
            return (sum(vals) / len(vals)) if vals else None
        if len(hd) >= 10:
            occ = avg(hd, "occ")
            if occ is None:                      # fill financials gap from CoStar
                occ = srcavg(subj_costar, "occ")
            out[label] = {"mkt": avg(hd, "mkt"), "eff": avg(hd, "eff"),
                          "occ": occ, "conc": avg(hd, "conc"), "src": "HD"}
        else:
            ask = srcavg(fb, "ask")              # rents from the better-fit source
            eff = srcavg(fb, "eff")
            occ = avg(hd, "occ")                 # occupancy: financials, else CoStar
            if occ is None:
                occ = srcavg(subj_costar, "occ")
            conc = avg(hd, "conc")
            if conc is None:
                conc = srcavg(subj_costar, "conc")
            if ask or eff or occ is not None:
                out[label] = {"conc": conc, "occ": occ, "src": fb_name,
                              "mkt": (ask * ratio) if ask else None,
                              "eff": (eff * ratio) if eff else None}
    return out, fb_name


def build_forecast_sheet(wb, series, props, as_of, target, latest_uc=None,
                         latest_label=None, subj_monthly=None, subj_costar=None,
                         subj_realpage=None, subject_name="", hist_years=6,
                         fwd_years=6, model_link=True,
                         model_sheet="Cash Flow (Annual)"):
    """Relative-year (trailing-12-month) supply / absorption / occupancy view.

    Columns are TTM windows: Y0 = the T12 ending at the as-of quarter, -Y1..-Yn
    step back a year each, Y1..Ym are the hold years anchored to an editable Close
    Quarter. Historical = CoStar actuals; forecast grows inventory by scheduled
    pipeline deliveries and occupied units by the Bear/Base/Bull annual demand.
    """
    ws = wb.create_sheet("Supply & Absorption")
    ws.sheet_view.showGridLines = False
    NAVYF = font(bold=True, size=9, color=WHITE)
    EDIT = fill("FFFFF2CC")

    as_idx = quarter_index(*as_of)
    close_idx = as_idx + 2                       # default close = +2 quarters
    cy, cq = close_idx // 4, close_idx % 4 + 1

    # ----- precompute historical annual (TTM) actuals from CoStar -----
    def win_sum(end_idx, key):
        tot = 0; any_ = False
        for k in range(4):
            rec = series.get(((end_idx - k) // 4, (end_idx - k) % 4 + 1))
            if rec and rec.get(key) is not None:
                tot += rec[key]; any_ = True
        return tot if any_ else None

    def at(end_idx, key):
        rec = series.get((end_idx // 4, end_idx % 4 + 1))
        return rec.get(key) if rec else None

    # relative-year column plan: list of (label, kind, k)
    plan = []
    for off in range(hist_years, -1, -1):          # -Y6 ... Y0
        plan.append((f"-Y{off}" if off else "Y0", "hist", off))
    for k in range(1, fwd_years + 1):              # Y1 ... Y6
        plan.append((f"Y{k}", "fcst", k))

    # ----- title -----
    ncols = len(plan)
    last_col = 2 + ncols                            # B=labels, C.. = years
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    t = ws.cell(2, 2, "SUPPLY & ABSORPTION  —  5-MILE MARKET (relative-year / TTM)")
    t.fill = fill(NAVY); t.font = font(bold=True, size=13, color=WHITE)
    t.alignment = CENTER
    ws.row_dimensions[2].height = 22

    # ----- assumptions (editable, yellow) -----
    base = round(hist_annual_absorption(series, as_of) / 25) * 25
    bear = max(0, round(base * 0.5 / 25) * 25)
    bull = round(base * 1.5 / 25) * 25
    A = [
        ("As-of Quarter", fmt_quarter(*as_of), None, False),
        ("Close Quarter (Y1 start)", fmt_quarter(cy, cq), None, True),
        ("Stabilization Target", target, "0%", True),
        ("Demand Scenario", "Base", None, True),
        ("Bear Annual Absorption", bear, "#,##0", True),
        ("Base Annual Absorption", base, "#,##0", True),
        ("Bull Annual Absorption", bull, "#,##0", True),
        ("Selected Annual Demand",
         '=IF($C$7="Bear",$C$8,IF($C$7="Bull",$C$10,$C$9))', "#,##0", False),
    ]
    for i, (lab, val, fmt, editable) in enumerate(A, start=4):
        ws.cell(i, 2, lab).font = font(size=9, bold=True)
        c = ws.cell(i, 3, val)
        c.font = font(size=9, bold=True); c.alignment = CENTER
        if fmt:
            c.number_format = fmt
        if editable:
            c.fill = EDIT
    dv = DataValidation(type="list", formula1='"Bear,Base,Bull"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(ws["C7"])
    # numeric helpers in hidden column Z (asof / close index / selected rank)
    ws["Z1"] = as_idx
    ws["Z2"] = ('=IFERROR(VALUE(RIGHT($C$5,4))*4+'
                'MATCH(LEFT($C$5,2),{"Q1","Q2","Q3","Q4"},0)-1,0)')
    ws["Z3"] = '=IF($C$7="Bear",3,IF($C$7="Base",2,1))'
    ws.column_dimensions["Z"].hidden = True
    DEMAND = "$C$11"; TARGET = "$C$6"
    CLOSE = "$Z$2"; ASOF = "$Z$1"; SELRANK = "$Z$3"

    # ----- pipeline blocks (cols R:X) -----
    # (A) UNDER CONSTRUCTION — certain supply, LINKED from the Competitive
    #     Analysis roster (units & delivery date), counts in every scenario.
    # (B) PROPOSED — speculative; per-deal "Built In" scenario toggle so you can
    #     layer a deal into the Bear (downside / more-supply) case only.
    uc = [p for p in props if p.bucket == "UNDER CONSTRUCTION" and p.roster_row]
    prop = [p for p in props if p.bucket == "PROPOSED"]

    def piped_headers(row, title, cols):
        ws.cell(row - 1, 18, title).font = NAVYF
        ws.cell(row - 1, 18).fill = fill(NAVY)
        for j, lab in enumerate(cols):
            cc = ws.cell(row, 18 + j, lab)
            cc.fill = fill(BLUE); cc.font = font(bold=True, size=8, color=WHITE)
            cc.alignment = CENTER; cc.border = BORDER

    # Visible cols R(Property) S(Units) T(Est Delivery) U(Built In, proposed only);
    # hidden helper cols V(DelivIdx) W(HoldYr) X(Built?).
    DIDX = ('=IFERROR(VALUE(RIGHT(T{r},4))*4+'
            'MATCH(LEFT(T{r},2),{{"Q1","Q2","Q3","Q4"}},0)-1,"")')
    HYR = ('=IF(V{r}="","",IF(V{r}<=' + ASOF + ',0,'
           'MAX(1,INT((V{r}-' + CLOSE + ')/4)+1)))')

    # (A) UC block
    uc_hdr = 4
    piped_headers(uc_hdr, "UNDER CONSTRUCTION (linked from Competitive Analysis)",
                  ["Property", "Units", "Est. Delivery"])
    rr = uc_hdr + 1
    for p in sorted(uc, key=lambda x: -(x.units or 0)):
        rw = p.roster_row
        ws.cell(rr, 18, f"='Competitive Analysis'!C{rw}").font = font(size=9)
        ws.cell(rr, 19, f"='Competitive Analysis'!D{rw}").number_format = "#,##0"
        ws.cell(rr, 20, f"='Competitive Analysis'!E{rw}")
        ws.cell(rr, 22, DIDX.format(r=rr))
        ws.cell(rr, 23, HYR.format(r=rr))
        for col in range(18, 21):
            cc = ws.cell(rr, col); cc.border = BORDER
            cc.alignment = LEFT if col == 18 else CENTER
            if col == 18:
                cc.font = font(size=9)
        rr += 1
    ucf, ucl = uc_hdr + 1, max(uc_hdr + 1, rr - 1)
    UC_U = f"$S${ucf}:$S${ucl}"; UC_H = f"$W${ucf}:$W${ucl}"

    # (B) Proposed block
    pp_hdr = ucl + 3
    piped_headers(pp_hdr, "PROPOSED PIPELINE (scenario toggle)",
                  ["Property", "Units", "Est. Delivery", "Built In"])
    builtin_dv = DataValidation(
        type="list", formula1='"Bear only,Bear+Base,All,None"', allow_blank=True)
    ws.add_data_validation(builtin_dv)
    rr = pp_hdr + 1
    for p in sorted(prop, key=lambda x: -(x.units or 0)):
        est = p.est_delivery if (p.deliv_year and p.deliv_q) else ""
        ws.cell(rr, 18, p.name).font = font(size=9)
        ws.cell(rr, 19, p.units).number_format = "#,##0"
        ws.cell(rr, 20, est)
        ws.cell(rr, 21, "Bear only")                  # default: downside case only
        ws.cell(rr, 22, DIDX.format(r=rr))
        ws.cell(rr, 23, HYR.format(r=rr))
        # Built? per selected scenario (more-supply/Bear has the most deals)
        ws.cell(rr, 24, f'=IF(U{rr}="All",1,IF(U{rr}="Bear+Base",'
                        f'IF({SELRANK}>=2,1,0),IF(U{rr}="Bear only",'
                        f'IF({SELRANK}>=3,1,0),0)))')
        for col in range(18, 22):
            cc = ws.cell(rr, col); cc.border = BORDER
            cc.alignment = LEFT if col == 18 else CENTER
            if col == 18:
                cc.font = font(size=9)
        for col in (20, 21):
            ws.cell(rr, col).fill = EDIT
        builtin_dv.add(ws.cell(rr, 21))
        rr += 1
    ppf, ppl = pp_hdr + 1, max(pp_hdr + 1, rr - 1)
    PP_U = f"$S${ppf}:$S${ppl}"; PP_H = f"$W${ppf}:$W${ppl}"; PP_B = f"$X${ppf}:$X${ppl}"
    for hcol in ("V", "W", "X"):
        ws.column_dimensions[hcol].hidden = True

    # ----- subject annual series (HelloData -> CoStar, level-aligned) -----
    subj_annual, fb_name = subject_annual_by_window(
        plan, as_idx, subj_monthly, subj_costar, subj_realpage)

    # ----- annual table -----
    HR = 17                                    # relative-year header row
    ws.cell(HR, 2, "5-MILE MARKET").font = font(bold=True, size=9)
    rows = {"period": HR + 1, "supply": HR + 2, "absorp": HR + 3, "occ": HR + 4,
            "smkt": HR + 6, "smkt_yoy": HR + 7, "seff": HR + 8, "seff_yoy": HR + 9,
            "socc": HR + 10, "sconc": HR + 11, "_inv": HR + 13, "_occu": HR + 14}
    ws.cell(rows["supply"], 2, "NEW SUPPLY (5-mi)").font = font(size=9, bold=True)
    ws.cell(rows["absorp"], 2, "ABSORPTION (5-mi)").font = font(size=9, bold=True)
    ws.cell(rows["occ"], 2, "OCCUPANCY (5-mi)").font = font(size=9, bold=True)
    ws.cell(HR + 5, 2, "SUBJECT (" + (subject_name or "subject") + ")").font = font(bold=True, size=9)
    ws.cell(rows["smkt"], 2, f"  Market Rent  ({fb_name}→HD)").font = font(size=9, bold=True)
    ws.cell(rows["smkt_yoy"], 2, "    Market Rent YoY %").font = font(size=8, italic=True)
    ws.cell(rows["seff"], 2, "  Effective Rent").font = font(size=9, bold=True)
    ws.cell(rows["seff_yoy"], 2, "    Effective Rent YoY %").font = font(size=8, italic=True)
    ws.cell(rows["socc"], 2, "  Occupancy (financials / CoStar)").font = font(size=9, bold=True)
    ws.cell(rows["sconc"], 2, "  Concession %").font = font(size=9, bold=True)
    ws.cell(rows["_inv"], 2, "  inventory").font = font(size=8, color="FF808080")
    ws.cell(rows["_occu"], 2, "  occupied").font = font(size=8, color="FF808080")

    def L(ci):
        return get_column_letter(ci)

    for j, (label, kind, k) in enumerate(plan):
        ci = 3 + j
        col = L(ci)
        # header band
        hc = ws.cell(HR, ci, label)
        hc.fill = fill(NAVY if kind == "hist" else BLUE)
        hc.font = NAVYF; hc.alignment = CENTER; hc.border = BORDER
        if kind == "hist":
            end_idx = as_idx - 4 * k
            sup = win_sum(end_idx, "deliveries")
            ab = win_sum(end_idx, "absorption")
            occ = at(end_idx, "occ_pct")
            inv = _int(at(end_idx, "inventory"))
            occu = _int(at(end_idx, "occ_units"))
            ey, eq = end_idx // 4, end_idx % 4 + 1
            sy, sq = (end_idx - 3) // 4, (end_idx - 3) % 4 + 1
            ws.cell(rows["period"], ci,
                    f"Q{sq}'{sy % 100:02d}-Q{eq}'{ey % 100:02d}")
            ws.cell(rows["supply"], ci, sup)
            ws.cell(rows["absorp"], ci, ab)
            ws.cell(rows["occ"], ci, occ)
            ws.cell(rows["_inv"], ci, inv)
            ws.cell(rows["_occu"], ci, occu)
            # subject rows (historical)
            srec = subj_annual.get(label)
            if srec:
                ws.cell(rows["smkt"], ci, round(srec["mkt"]) if srec.get("mkt") else None)
                ws.cell(rows["seff"], ci, round(srec["eff"]) if srec.get("eff") else None)
                ws.cell(rows["socc"], ci, srec.get("occ"))
        else:
            prev = L(ci - 1)
            ws.cell(rows["period"], ci, f"Hold Yr {k}")
            # New supply = UC (always) + proposed that are "built" this scenario
            ws.cell(rows["supply"], ci,
                    f"=SUMIFS({UC_U},{UC_H},{k})"
                    f"+SUMIFS({PP_U},{PP_H},{k},{PP_B},1)")
            ws.cell(rows["_inv"], ci,
                    f"={prev}{rows['_inv']}+{col}{rows['supply']}")
            ws.cell(rows["_occu"], ci,
                    f"=MIN({TARGET}*{col}{rows['_inv']},"
                    f"{prev}{rows['_occu']}+{DEMAND})")
            ws.cell(rows["absorp"], ci,
                    f"={col}{rows['_occu']}-{prev}{rows['_occu']}")
            ws.cell(rows["occ"], ci,
                    f"=IFERROR({col}{rows['_occu']}/{col}{rows['_inv']},0)")
        # formatting
        for rk in ("supply", "absorp", "_inv", "_occu"):
            cc = ws.cell(rows[rk], ci); cc.number_format = "#,##0"
            cc.alignment = CENTER; cc.font = font(size=9); cc.border = BORDER
        for rk in ("occ", "socc", "sconc"):
            oc = ws.cell(rows[rk], ci)
            oc.number_format = "0.0%"; oc.alignment = CENTER
            oc.font = font(size=9, bold=(rk == "occ")); oc.border = BORDER
        for rk in ("smkt_yoy", "seff_yoy"):
            yc = ws.cell(rows[rk], ci)
            yc.number_format = "0.0%"; yc.alignment = CENTER
            yc.font = font(size=8, italic=True, color="FF808080"); yc.border = BORDER
        for rk in ("smkt", "seff"):
            sc = ws.cell(rows[rk], ci)
            sc.number_format = '"$"#,##0'; sc.alignment = CENTER
            sc.font = font(size=9); sc.border = BORDER
        pc = ws.cell(rows["period"], ci)
        pc.font = font(size=8, color="FF808080"); pc.alignment = CENTER
        band = fill("FFF2F2F2" if kind == "hist" else "FFFFFFFF")
        for rk in ("supply", "absorp", "occ", "smkt", "smkt_yoy",
                   "seff", "seff_yoy", "socc", "sconc"):
            ws.cell(rows[rk], ci).fill = band
    # hide helper rows
    for rk in ("_inv", "_occu"):
        ws.row_dimensions[rows[rk]].hidden = True

    # ----- link the subject rows to the underwriting model (default) -----
    # Internal references to the TMG model's "Cash Flow (Annual)" tab, so when
    # these tabs are dragged into (or embedded in) the model they resolve to its
    # own projection: Y0 <- col F (T12/Y0), Y1..Y6 <- cols K..P; Market Rent <-
    # row 4, Effective Rent <- row 5, Occupancy <- row 14. Wrapped in IFERROR so
    # the standalone file shows a clean blank (Y1..Y6) or the chart's own value
    # (Y0 fallback) until the tab is carried into the model. Historical columns
    # (-Y6..-Y1) stay as the static RealPage/HelloData actuals.
    if model_link:
        link_rows = {"smkt": 4, "seff": 5, "socc": 14}
        chart_cols = [3 + hist_years + i for i in range(fwd_years + 1)]   # Y0,Y1..Y6
        model_cols = ["F"] + [get_column_letter(11 + k) for k in range(fwd_years)]  # F,K..P
        for rk, mrow in link_rows.items():
            for idx, (ci, mcol) in enumerate(zip(chart_cols, model_cols)):
                cell = ws.cell(rows[rk], ci)
                fb = cell.value if (idx == 0 and isinstance(cell.value, (int, float))) else None
                link = f"'{model_sheet}'!${mcol}${mrow}"
                cell.value = (f"=IFERROR({link},{fb})" if fb is not None
                              else f'=IFERROR({link},"")')

    # ----- subject derived rows (YoY % + concession %) — all columns -----
    # Written for every year (hist + forecast) so they populate wherever the
    # subject rent rows have data (static history, or the model links once in
    # the model). IFERROR keeps them blank when a year has no rent.
    for j in range(len(plan)):
        ci = 3 + j
        col = L(ci)
        ws.cell(rows["sconc"], ci,
                f'=IFERROR(({col}{rows["smkt"]}-{col}{rows["seff"]})/{col}{rows["smkt"]},"")')
        if j > 0:
            pcl = L(ci - 1)
            ws.cell(rows["smkt_yoy"], ci,
                    f'=IFERROR({col}{rows["smkt"]}/{pcl}{rows["smkt"]}-1,"")')
            ws.cell(rows["seff_yoy"], ci,
                    f'=IFERROR({col}{rows["seff"]}/{pcl}{rows["seff"]}-1,"")')

    # ----- scenario block: implied 5-mi occupancy (Bear/Base/Bull) -----
    y0_occu = f"${L(3 + hist_years)}${rows['_occu']}"   # actual occupied at Y0
    sb = rows["_occu"] + 2          # below the hidden inventory/occupied helpers
    ws.cell(sb, 2, "IMPLIED 5-MI OCCUPANCY BY DEMAND SCENARIO").font = font(bold=True, size=9)
    fcst_cols = [3 + j for j, (_, kind, _) in enumerate(plan) if kind == "fcst"]
    for r_off, (sc_lab, abs_cell) in enumerate(
            [("Bear", "$C$8"), ("Base", "$C$9"), ("Bull", "$C$10")], start=1):
        rrow = sb + r_off
        lc = ws.cell(rrow, 2, sc_lab); lc.font = font(size=9, bold=True)
        for k, ci in enumerate(fcst_cols, start=1):
            col = L(ci)
            cc = ws.cell(rrow, ci,
                         f"=MIN({TARGET},({y0_occu}+{abs_cell}*{k})/{col}{rows['_inv']})")
            cc.number_format = "0.0%"; cc.alignment = CENTER; cc.font = font(size=9)
            cc.border = BORDER
        # label the forecast year columns above
        if r_off == 1:
            for k, ci in enumerate(fcst_cols, start=1):
                hc = ws.cell(sb, ci, f"Y{k}")
                hc.font = NAVYF; hc.fill = fill(BLUE); hc.alignment = CENTER

    # ----- rent-growth scenarios (editable assumptions) -----
    rg = sb + 5
    ws.cell(rg, 2, "MARKET RENT GROWTH (editable assumptions)").font = font(bold=True, size=9)
    base_g = [0.01, 0.03, 0.04, 0.05, 0.04, 0.03]
    conc_b = [0.03, 0.015, 0.005, 0.005, 0.005, 0.005]
    for k, ci in enumerate(fcst_cols, start=1):
        ws.cell(rg, ci, f"Y{k}").font = font(bold=True, size=8)
    growth_rows = {}
    for r_off, sc in enumerate(["Bear", "Base", "Bull"], start=1):
        rrow = rg + r_off
        ws.cell(rrow, 2, sc).font = font(size=9, bold=True)
        growth_rows[sc] = rrow
        for k, ci in enumerate(fcst_cols):
            g = base_g[k] + (-0.0125 if sc == "Bear" else 0.01 if sc == "Bull" else 0)
            cc = ws.cell(rrow, ci, round(g, 4))
            cc.number_format = "0.0%"; cc.alignment = CENTER; cc.font = font(size=9)
            cc.fill = EDIT; cc.border = BORDER

    cg = rg + 4
    ws.cell(cg, 2, "CONCESSIONS % — forward assumption (editable)").font = font(bold=True, size=9)
    conc_rows = {}
    for r_off, sc in enumerate(["Bear", "Base", "Bull"], start=1):
        rrow = cg + r_off
        ws.cell(rrow, 2, sc).font = font(size=9, bold=True)
        conc_rows[sc] = rrow
        for k, ci in enumerate(fcst_cols):
            c = conc_b[k] + (0.03 if sc == "Bear" else -0.01 if sc == "Bull" else 0)
            c = max(c, 0)
            cc = ws.cell(rrow, ci, round(c, 4))
            cc.number_format = "0.0%"; cc.alignment = CENTER; cc.font = font(size=9)
            cc.fill = EDIT; cc.border = BORDER

    eg = cg + 4
    ws.cell(eg, 2, "EFFECTIVE RENT GROWTH (derived)").font = font(bold=True, size=9)
    for r_off, sc in enumerate(["Bear", "Base", "Bull"], start=1):
        rrow = eg + r_off
        ws.cell(rrow, 2, sc).font = font(size=9, bold=True)
        for k, ci in enumerate(fcst_cols):
            col = L(ci)
            gr = f"{col}{growth_rows[sc]}"
            ct = f"{col}{conc_rows[sc]}"
            cp = f"{L(ci-1)}{conc_rows[sc]}" if k > 0 else None
            if k == 0:
                f = f"=(1+{gr})*(1-{ct})-1"
            else:
                f = f"=(1+{gr})*(1-{ct})/(1-{cp})-1"
            cc = ws.cell(rrow, ci, f)
            cc.number_format = "0.0%"; cc.alignment = CENTER; cc.font = font(size=9)
            cc.border = BORDER

    # ----- collapsed: subject rent source by year + reconciliation -----
    sg = eg + 5
    ws.cell(sg, 2, "▸ DETAIL: subject rent source by year & reconciliation "
                   "(grouped — click − to collapse)").font = font(bold=True, size=8)
    grp_first = sg + 1
    r = grp_first
    for (label, kind, k) in plan:
        if kind != "hist":
            continue
        srec = subj_annual.get(label)
        src = srec.get("src") if srec else "—"
        end_idx = as_idx - 4 * k
        ws.cell(r, 2, f"  {label}  "
                f"{fmt_quarter((end_idx-3)//4,(end_idx-3)%4+1)}–"
                f"{fmt_quarter(end_idx//4,end_idx%4+1)}").font = font(size=8)
        ws.cell(r, 5, src).font = font(size=8)
        r += 1
    uc_total = sum(p.units or 0 for p in uc)
    ws.cell(r, 2, "  UC scheduled vs CoStar UC").font = font(size=8)
    ws.cell(r, 5, f"{uc_total:,} / {latest_uc:,}" if latest_uc else f"{uc_total:,}").font = font(size=8)
    r += 1
    if latest_label and "QTD" in str(latest_label):
        ws.cell(r, 2, f"  As-of {latest_label} is partial (QTD); occupancy is "
                      f"point-in-time, Y0 flow sums partial.").font = font(size=8, color="FF808080")
        r += 1
    for gr in range(grp_first, r):
        ws.row_dimensions[gr].outlineLevel = 1
        ws.row_dimensions[gr].hidden = True
    ws.sheet_properties.outlinePr.summaryBelow = False

    # ----- column widths + compact row heights -----
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 26
    for ci in range(3, last_col + 1):
        ws.column_dimensions[L(ci)].width = 9.5
    for cw, w in {"R": 28, "S": 7, "T": 11, "U": 9, "V": 8, "W": 7, "X": 6}.items():
        ws.column_dimensions[cw].width = w
    for rr2 in range(1, r + 2):
        ws.row_dimensions[rr2].height = 14.4
    ws.row_dimensions[2].height = 20            # title
    return ws

def write_workbook(props, subject_name, latest_inv, latest_label,
                   as_of, target, out_path, series=None, latest_uc=None,
                   subj_monthly=None, subj_costar=None, subj_realpage=None,
                   diligence_rows=None, model_link=True,
                   model_sheet="Cash Flow (Annual)"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competitive Analysis"
    ws.sheet_view.showGridLines = False

    for col, (_, width) in COLS.items():
        ws.column_dimensions[col].width = width
    ws.column_dimensions["A"].width = 3

    # ---- Title / subtitle ----
    ws.merge_cells("C2:G2")
    t = ws["C2"]
    t.value = f"{subject_name.upper()} — 5-Mile Competitive Supply"
    t.fill = fill(NAVY); t.font = font(bold=True, size=13, color=WHITE)
    t.alignment = CENTER
    ws.row_dimensions[2].height = 22

    delivered_years = sorted({p.deliv_year for p in props if p.deliv_year})
    yr_lo = delivered_years[0] if delivered_years else as_of[0]
    ws.merge_cells("B3:F3")
    s = ws["B3"]
    s.value = (f"{yr_lo}–{as_of[0]} New Construction  |  5-mile radius  |  "
               f"As of {latest_label}  |  Stabilization Target:")
    s.fill = fill(BLUE); s.font = font(size=9, color=WHITE); s.alignment = LEFT
    g = ws["G3"]
    g.value = f"{target:.0%} target"
    g.fill = fill(BLUE); g.font = font(bold=True, size=9, color=YELLOW)
    g.alignment = CENTER

    # ---- Header row 4 ----
    hr = 4
    for col, (label, _) in COLS.items():
        c = ws[f"{col}{hr}"]
        c.value = label
        c.fill = fill(NAVY); c.font = font(bold=True, size=9, color=WHITE)
        c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[hr].height = 18

    # ---- Buckets ----
    r = hr + 1
    idx = 0
    for bucket, _desc in BUCKET_ORDER:
        members = [p for p in props if p.bucket == bucket]
        if not members:
            continue
        # chronological within each section: earliest delivery first, undated last
        members.sort(key=lambda p: (p_qi(p) if (p.deliv_year and p.deliv_q) else 10**9,
                                    -(p.units or 0)))
        tot_u = sum(p.units or 0 for p in members)
        hdr_color = BUCKET_STYLES[bucket]["header"]
        row_color = BUCKET_STYLES[bucket]["row"]
        # section header
        ws.merge_cells(f"C{r}:K{r}")
        sc = ws[f"C{r}"]
        sc.value = f"{bucket} ({len(members)} properties, {tot_u:,} units)"
        sc.fill = fill(hdr_color); sc.font = font(bold=True, size=9, color=WHITE)
        sc.alignment = LEFT
        ws[f"B{r}"].fill = fill(hdr_color)
        r += 1
        for p in members:
            idx += 1
            p.roster_row = r          # remember for cross-sheet links
            row_vals = {
                "B": idx,
                "C": p.name,
                "D": p.units,
                "E": p.est_delivery or "TBD",
                "F": p.occupancy,
                "G": p.asking_rent if p.asking_rent else "—",
                "H": p.owner or "—",
                "I": (round(p.proximity_mi, 1)
                      if p.proximity_mi is not None else None),
                "J": p.prop_type,
                "K": "; ".join(p.notes) if p.notes else None,
            }
            for col, val in row_vals.items():
                c = ws[f"{col}{r}"]
                c.value = val
                c.fill = fill(row_color); c.font = font(size=9)
                c.border = BORDER
                c.alignment = LEFT if col in ("C", "H", "K") else CENTER
            ws[f"D{r}"].number_format = "#,##0"
            ws[f"F{r}"].number_format = "0.0%"
            ws[f"G{r}"].number_format = '"$"#,##0;;"—"'
            ws[f"I{r}"].number_format = '0.0" mi";;"—"'
            r += 1
        r += 1  # spacer

    # ---- Pointer note (forward forecast lives on its own tab) ----
    note_r = r + 1
    ws.merge_cells(f"B{note_r}:K{note_r}")
    ws[f"B{note_r}"] = ("* 5-Mile radius. Rent ($) = market asking. Proximity (mi) = "
                        "straight-line distance from the subject (geocoded); lease-up "
                        "rent/occupancy to be verified w/ HelloData. See the 'Supply & "
                        "Absorption' tab for the forward supply / absorption / "
                        "overall-occupancy forecast.")
    ws[f"B{note_r}"].font = font(size=8, color="FF808080")

    # ---- Reconciliation log sheet ----
    rlog = wb.create_sheet("Reconciliation Log")
    rlog.append(["Property", "Address", "Units", "Year Built", "Est. Delivery",
                 "Bucket", "CoStar Occ", "RealPage Occ",
                 "CoStar Ask Rent", "RealPage Eff Rent",
                 "Sources", "Notes"])
    for c in rlog[1]:
        c.font = font(bold=True, color=WHITE); c.fill = fill(NAVY)
        c.alignment = CENTER
    for p in sorted(props, key=lambda x: (BUCKET_ORDER.index(
            next(b for b in BUCKET_ORDER if b[0] == x.bucket)),
            -p_qi(x))):
        rlog.append([p.name, p.address, p.units, p.year_built,
                     p.est_delivery, p.bucket,
                     p.costar_occ, p.rp_occ, p.costar_rent, p.rp_rent,
                     "+".join(sorted(p.sources)), "; ".join(p.notes)])
    for i, p in enumerate(sorted(props, key=lambda x: (BUCKET_ORDER.index(
            next(b for b in BUCKET_ORDER if b[0] == x.bucket)), -p_qi(x))), 2):
        rlog[f"G{i}"].number_format = "0.0%"
        rlog[f"H{i}"].number_format = "0.0%"
        rlog[f"I{i}"].number_format = "#,##0"
        rlog[f"J{i}"].number_format = "#,##0"
    for col in "ABCDEFGHIJKL":
        rlog.column_dimensions[col].width = 15
    rlog.column_dimensions["A"].width = 30
    rlog.column_dimensions["B"].width = 24
    rlog.column_dimensions["L"].width = 44

    # ---- Forward supply / absorption / occupancy forecast ----
    if series:
        sa = build_forecast_sheet(wb, series, props, as_of, target, latest_uc=latest_uc,
                                  latest_label=latest_label, subj_monthly=subj_monthly,
                                  subj_costar=subj_costar, subj_realpage=subj_realpage,
                                  subject_name=subject_name, model_link=model_link,
                                  model_sheet=model_sheet)
        # make "Supply & Absorption" the 2nd tab
        wb.move_sheet(sa, -(wb.index(sa) - 1))

    if diligence_rows:
        write_diligence_sheet(wb, diligence_rows)

    wb.save(out_path)


def write_diligence_sheet(wb, rows):
    """Render researched per-project diligence + a shadow-supply watch list."""
    ws = wb.create_sheet("Diligence")
    ws.sheet_view.showGridLines = False
    headers = ["Property / Site", "Units", "Est. Delivery", "Status",
               "Leasing Pace", "Notes", "Source"]
    widths = [30, 8, 12, 16, 14, 50, 36]
    for sect, kind in (("PIPELINE DILIGENCE", "pipeline"),
                       ("SHADOW SUPPLY (latent / untracked — watch list)", "shadow")):
        srows = [r for r in rows if r.get("type", "pipeline") == kind]
        if not srows:
            continue
        r0 = (ws.max_row + 2) if ws.max_row > 1 else 2
        ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=8)
        sc = ws.cell(r0, 2, sect)
        sc.fill = fill(NAVY); sc.font = font(bold=True, size=11, color=WHITE)
        for j, h in enumerate(headers):
            c = ws.cell(r0 + 1, 2 + j, h)
            c.fill = fill(BLUE); c.font = font(bold=True, size=9, color=WHITE)
            c.alignment = CENTER; c.border = BORDER
        for i, r in enumerate(srows):
            rr = r0 + 2 + i
            vals = [r.get("property"), r.get("units"), r.get("est_delivery"),
                    r.get("status"), r.get("leasing_pace"), r.get("notes"),
                    r.get("source")]
            for j, v in enumerate(vals):
                c = ws.cell(rr, 2 + j, v or None)
                c.font = font(size=9); c.border = BORDER
                c.alignment = LEFT if j in (0, 3, 4, 5, 6) else CENTER
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + j)].width = w
    ws.column_dimensions["A"].width = 2.5
    return ws



def p_qi(p):
    return quarter_index(p.deliv_year, p.deliv_q) if p.deliv_year and p.deliv_q else -1


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def parse_as_of(label: str) -> tuple[int, int]:
    m = _QRE.search(label or "")
    if m:
        return int(m.group(1)), int(m.group(2))
    today = dt.date.today()
    return today.year, (today.month - 1) // 3 + 1


def build_competitive_roster(costar_roster_path, realpage_path, deliveries, as_of,
                             subject_name, subject_address, target=0.95,
                             occ_source="costar", rent_source="costar"):
    """Parse + reconcile + classify the competitive new-construction roster.

    Shared by the chart and the map. Returns the kept Prop list (subject excluded,
    bucketed, with delivery quarters pinned/estimated)."""
    props = reconcile(parse_costar_roster(costar_roster_path),
                      parse_realpage(realpage_path))
    subj_addr = addr_key(subject_address)
    subj_name = name_key(subject_name)
    props = [p for p in props
             if not ((subj_addr and addr_key(p.address) == subj_addr)
                     or name_key(p.name) == subj_name)]
    keep = []
    for p in props:
        delivered = _existing_signal(p)
        if delivered:
            pin_delivery(p, deliveries)
        else:
            estimate_pipeline_delivery(p, as_of)
        resolve_occ_rent(p, occ_source, rent_source, flag_divergence=delivered)
        classify(p, as_of, target)
        p.prop_type = derive_type(p)
        within_lookback = (p.year_built and
                           p.year_built >= as_of[0] - NEW_CONSTRUCTION_LOOKBACK_YEARS)
        if _is_pipeline(p) or within_lookback:
            keep.append(p)
    return keep


def reconcile_deliveries(props, series, as_of, lookback=4):
    """Print delivered-units reconciliation vs CoStar's analytics series per TTM.

    For each historical trailing-12-month window, compares the named roster comps'
    delivered units to CoStar's submarket `deliveries`. They should tie; a positive
    gap (CoStar > roster) is normally the **subject** (counted in the submarket
    total but excluded from the competitive roster) plus any sub-50-unit product.
    A gap that those don't explain means a 50+ comp is missing or mis-dated — the
    case worth catching (see SKILL.md §3a)."""
    if not series:
        return
    aq = quarter_index(*as_of)
    ql = lambda i: f"Q{i % 4 + 1} {i // 4}"
    rows = []
    for k in range(lookback, -1, -1):
        end, start = aq - 4 * k, aq - 4 * k - 3
        costar = sum((v.get("deliveries") or 0) for (y, q), v in series.items()
                     if start <= quarter_index(y, q) <= end)
        members = [p for p in props if p.deliv_year
                   and start <= quarter_index(p.deliv_year, p.deliv_q) <= end]
        roster = sum(p.units or 0 for p in members)
        rows.append((k, start, end, costar, roster, members))
    if not any(r[3] or r[4] for r in rows):
        return
    print("\n  Delivered-units reconciliation vs CoStar analytics (TTM windows):")
    flagged = False
    for k, start, end, costar, roster, members in rows:
        diff = costar - roster
        flag = ""
        if diff > 0:
            flag = f"  (+{diff:,.0f} = subject / sub-50 product?)"
        elif diff < 0:
            flag = f"  (roster exceeds CoStar by {-diff:,.0f} — check pins)"; flagged = True
        print(f"    -Y{k} [{ql(start)}..{ql(end)}]  CoStar {costar:>6,.0f}  |  "
              f"roster {roster:>6,}{flag}")
        if abs(diff) >= 50 and diff > 0:
            flagged = True
    if flagged:
        print("    ⚠ Reconcile the flagged window(s): add the subject's units "
              "back, account for sub-50 product, else look for a missing/mis-dated "
              "50+ comp (SKILL.md §3a).")


def compute_proximity(props, subject_name, subject_address, cache_path,
                      use_geocoder=True):
    """Set p.proximity_mi = straight-line miles from the subject to each comp.

    Uses the shared geo.Locator (Nominatim, cached) with an offline ZIP-centroid
    fallback. The subject must geocode to a real point; if it can't (no address /
    geocoding off and no ZIP), proximity is left as None for the analyst."""
    loc = geo.Locator(use_geocoder, cache_path)
    subj_ll, subj_approx = loc.locate(subject_name, subject_address, None, None,
                                      geo.subject_zip(subject_address))
    if not subj_ll:
        loc.save()
        return
    loc.set_anchor(subj_ll)            # reject comps mis-geocoded outside the radius
    for p in props:
        ll, _ = loc.locate(p.name, p.address, p.city, p.state, p.zipcode)
        if ll:
            p.proximity_mi = geo.haversine_miles(subj_ll, ll)
    loc.save()


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--subject-name", required=True)
    ap.add_argument("--subject-address", default="")
    ap.add_argument("--costar-roster", required=True)
    ap.add_argument("--costar-analytics", required=True)
    ap.add_argument("--realpage", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--as-of", default=None,
                    help="Override analysis quarter, e.g. '2026 Q2'")
    ap.add_argument("--target", type=float, default=DEFAULT_STABILIZATION_TARGET)
    ap.add_argument("--occ-source", choices=("costar", "realpage"), default="costar",
                    help="Which source's occupancy to display (default: costar)")
    ap.add_argument("--rent-source", choices=("costar", "realpage", "average"),
                    default="costar",
                    help="Source for the roster's Avg Mkt Rent (asking/market): "
                         "costar = CoStar asking (default); realpage = RealPage "
                         "asking IF its export carries one (else CoStar); average = "
                         "blend the available asking values. RealPage *effective* "
                         "rent is never used for this column.")
    ap.add_argument("--pipeline-dates", default=None,
                    help="CSV (property,est_delivery[,units]) of analyst-supplied "
                         "delivery quarters for pipeline deals, so they flow into "
                         "the absorption forecast.")
    ap.add_argument("--intake", default=None,
                    help="RR-T12 underwriting intake .xlsx (subject HelloData rents "
                         "+ financial-statement occupancy for the subject rows).")
    ap.add_argument("--costar-subject-rents", default=None,
                    help="CoStar per-property (50-unit) analytics export, used for "
                         "subject rent history before HelloData coverage begins.")
    ap.add_argument("--realpage-subject-rents", default=None,
                    help="RealPage per-property 10-yr rent export (alternative "
                         "subject rent source; the better-tracking of CoStar/"
                         "RealPage vs HelloData is auto-selected).")
    ap.add_argument("--diligence", default=None,
                    help="Filled per-project research CSV (see the emitted "
                         "…__diligence_TEMPLATE.csv) — adds a Diligence + shadow-"
                         "supply sheet and folds researched delivery dates back in.")
    ap.add_argument("--no-geocode", action="store_true",
                    help="Skip Nominatim geocoding for the Proximity column "
                         "(falls back to offline ZIP centroids only).")
    ap.add_argument("--no-model-link", action="store_true",
                    help="Don't wire the subject rows to the underwriting model. "
                         "By default the subject Market Rent / Effective Rent / "
                         "Occupancy rows (Y0..Y6) carry IFERROR'd internal refs to "
                         "'Cash Flow (Annual)' rows 4/5/14 (Y0<-F, Y1..Y6<-K:P), so "
                         "the tab is ready to drag into / embed in the model and "
                         "shows clean blanks until then. Pass this for a pure "
                         "standalone chart with no model references.")
    ap.add_argument("--model-sheet", default="Cash Flow (Annual)",
                    help="Model tab the subject-row refs point to "
                         "(default: 'Cash Flow (Annual)').")
    args = ap.parse_args(argv)

    latest_inv, deliveries, latest_label, series, latest_uc = parse_costar_analytics(args.costar_analytics)
    as_of = parse_as_of(args.as_of or latest_label)

    props = build_competitive_roster(
        args.costar_roster, args.realpage, deliveries, as_of,
        args.subject_name, args.subject_address, args.target,
        args.occ_source, args.rent_source)

    # Apply analyst-supplied pipeline delivery quarters so they enter the forecast.
    if args.pipeline_dates:
        applied = load_pipeline_dates(args.pipeline_dates)
        apply_pipeline_dates(props, applied)

    diligence_rows = None
    if args.diligence:
        diligence_rows = load_diligence(args.diligence)
        apply_diligence(props, diligence_rows)

    # Proximity: straight-line miles from the subject to each comp (geocoded,
    # cached). Falls back to offline ZIP centroids when geocoding is unavailable.
    import os
    compute_proximity(props, args.subject_name, args.subject_address,
                      os.path.join(os.path.dirname(args.out) or ".", ".geocode_cache.json"),
                      use_geocoder=not args.no_geocode)

    subj_monthly = parse_intake_subject(args.intake) if args.intake else {}
    subj_costar = (parse_costar_subject_rents(args.costar_subject_rents, args.subject_name)
                   if args.costar_subject_rents else {})
    subj_realpage = (parse_realpage_subject_rents(args.realpage_subject_rents, args.subject_name)
                     if args.realpage_subject_rents else {})

    write_workbook(props, args.subject_name, latest_inv, latest_label,
                   as_of, args.target, args.out, series=series, latest_uc=latest_uc,
                   subj_monthly=subj_monthly, subj_costar=subj_costar,
                   subj_realpage=subj_realpage, diligence_rows=diligence_rows,
                   model_link=not args.no_model_link, model_sheet=args.model_sheet)

    # Emit templates: undated pipeline + a per-project research template.
    base = os.path.splitext(args.out)[0]
    n_undated = emit_pipeline_template(props, base + "__pipeline_dates_TEMPLATE.csv")
    emit_diligence_template(props, base + "__diligence_TEMPLATE.csv")

    # Console reconciliation report
    print(f"\nSupply chart written: {args.out}")
    print(f"As-of quarter: {fmt_quarter(*as_of)}  |  "
          f"Total current inventory (CoStar): {latest_inv:,}")
    print(f"Competitive new-construction properties: {len(props)}\n")
    for bucket, _ in BUCKET_ORDER:
        members = [p for p in props if p.bucket == bucket]
        if not members:
            continue
        u = sum(p.units or 0 for p in members)
        print(f"  {bucket} — {len(members)} props, {u:,} units")
        for p in sorted(members, key=lambda x: -p_qi(x)):
            occ = f"{p.occupancy:.1%}" if p.occupancy is not None else "  —  "
            print(f"      {p.name[:34]:34s} {str(p.units or '—'):>5} u  "
                  f"{p.est_delivery or 'TBD':>8}  occ {occ:>6}  "
                  f"[{'+'.join(sorted(p.sources))}]"
                  f"{('  | ' + '; '.join(p.notes)) if p.notes else ''}")
    reconcile_deliveries(props, series, as_of)
    if n_undated:
        print(f"\n  {n_undated} pipeline deal(s) have no delivery quarter and are "
              f"NOT in the absorption forecast yet.\n"
              f"  Fill {base}__pipeline_dates_TEMPLATE.csv (--pipeline-dates), or "
              f"research them via {base}__diligence_TEMPLATE.csv (--diligence).")
    print()


if __name__ == "__main__":
    sys.exit(main())
