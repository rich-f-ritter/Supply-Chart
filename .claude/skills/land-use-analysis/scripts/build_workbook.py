"""Build the comprehensive, self-documenting Excel workbook - the canonical deliverable.
A reader should understand the ENTIRE analysis from this file alone AND audit the actual data.
Sheets:
  Overview · Assumptions & Decisions · Methodology · Data Sources · Land-Use Grouping ·
  Zoning Grouping · MF Threat Classification · Parcels (every classified parcel: land use +
  zoning + threat) · Vacant - Developable (the flagged threat set) · Top-N Vacant Threats.
Pulls from config.json + Tables/*.csv (classify/finalize) + in/parcels_classified.geojson +
in/vacant_candidates.json + in/developable_accounts.json + an optional Tables/decisions_log.md.
"""
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config as C  # noqa: E402
import palette as P  # noqa: E402

SERIF = "Cambria"
THREAT_FILL = {"High": "F4C7C3", "Medium": "FCE3C6", "Low": "FFF7D6", "Unknown": "E4D7F0"}


def _csv(path):
    return list(csv.DictReader(open(path, encoding="utf-8"))) if path.exists() else []


def threat_word(s):
    return (str(s or "").split("-")[0].split()[0].strip().capitalize()
            if s else "")


def main():
    cfg = C.load()
    TBL = C.tables(cfg)
    IN = C.indir(cfg)
    brand = cfg.get("branding", {})
    NAVY = brand.get("primary", "#074070").lstrip("#")
    GRAY, INK, BAND = "6F6F70", "1A2B3C", "F4F8FB"
    thin = Side(style="thin", color="D9D9D9")
    B = Border(thin, thin, thin, thin)
    name = cfg["subject"]["name"]

    wb = Workbook()
    wb.remove(wb.active)

    def title(ws, text, sub, ncols):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text); c.font = Font(name=SERIF, size=18, bold=True, color=NAVY)
        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(2, 1, sub).font = Font(name="Calibri", size=10, italic=True, color=GRAY)

    def header(ws, row, cols):
        for j, h in enumerate(cols, 1):
            c = ws.cell(row, j, h)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.alignment = Alignment("center", "center", wrap_text=True); c.border = B
        ws.row_dimensions[row].height = 28

    def body(ws, start, rows, widths, wrapcols=(), threatcol=None, catcolorcol=None, catcolors=None):
        for i, r in enumerate(rows):
            rr = start + i
            for j, v in enumerate(r, 1):
                c = ws.cell(rr, j, v); c.border = B
                c.font = Font(name="Calibri", size=10, color=INK)
                c.alignment = Alignment(vertical="top", wrap_text=(j - 1 in wrapcols))
                if i % 2:
                    c.fill = PatternFill("solid", fgColor=BAND)
                if threatcol is not None and j - 1 == threatcol:
                    tw = threat_word(v)
                    if tw in THREAT_FILL:
                        c.fill = PatternFill("solid", fgColor=THREAT_FILL[tw])
                        c.font = Font(name="Calibri", size=10, bold=True, color=INK)
                if catcolorcol is not None and j - 1 == catcolorcol and catcolors and catcolors[i]:
                    c.fill = PatternFill("solid", fgColor=catcolors[i].lstrip("#"))
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(start, 1)
        return start + len(rows)

    def paragraphs(ws, start, text, ncols=1, width=110):
        """Render markdown-ish text as wrapped rows (one block per blank-line group)."""
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = width
        r = start
        for line in text.splitlines():
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            cell = ws.cell(r, 1, line.rstrip())
            bold = line.startswith("#") or re.match(r"^\s*\d+\.\s", line) or line.startswith("**")
            txt = re.sub(r"^#+\s*", "", line).replace("**", "")
            cell.value = txt
            cell.font = Font(name="Calibri", size=11 if line.startswith("# ") else 10,
                             bold=bool(bold), color=NAVY if line.startswith("#") else INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = max(15, 15 * (1 + len(txt) // width))
            r += 1
        return r

    def lean_sheet(sheetname, title_text, sub, cols, rows, widths, threatcol=None):
        """Fast writer for large DATA sheets: styled header, values-only body (no per-cell
        styling so 20k+ rows stay quick/small), autofilter, freeze, and threat color via
        conditional formatting."""
        ws = wb.create_sheet(sheetname)
        title(ws, title_text, sub, len(cols))
        header(ws, 4, cols)
        for r in rows:
            ws.append(list(r))
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        last = 4 + len(rows)
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{max(last, 4)}"
        if threatcol is not None and rows:
            col = get_column_letter(threatcol + 1)
            rng = f"{col}5:{col}{last}"
            for tw, fill in THREAT_FILL.items():
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="equal", formula=[f'"{tw}"'],
                    fill=PatternFill("solid", fgColor=fill)))
        return ws

    # ── 1) Overview ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Overview")
    title(ws, f"{name} — Land Use Analysis",
          "Competitive land-use / zoning / multifamily-supply-threat analysis", 2)
    a = cfg["analysis_area"]
    area_desc = (f"{a.get('radius_mi')}-mile radius" if a.get("mode") != "polygon"
                 else f"custom polygon ({a.get('note', 'hand-drawn area')})")
    headline = (TBL / "headline.txt").read_text(encoding="utf-8") if (TBL / "headline.txt").exists() else \
        cfg.get("headline", "(headline pending reasoned threat assessment)")
    sub = cfg["subject"]
    info = [
        ("Subject", name),
        ("Address", sub.get("address", "—")),
        ("Verified point", f"{sub.get('lat')}, {sub.get('lon')}"),
        ("Location note", sub.get("location_note", "—")),
        ("Analysis area", area_desc),
        ("Vacant definition", cfg.get("landuse_scheme", {}).get("vacant_note",
                              ", ".join(cfg.get("landuse_scheme", {}).get("vacant_buckets", ["Vacant Land"])))),
        ("Min. developable acreage", cfg.get("min_acres", 1.0)),
        ("Compactness floor (Polsby-Popper)", cfg.get("min_compactness", 0.16)),
    ]
    r = 4
    for k, v in info:
        ws.cell(r, 1, k).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        c = ws.cell(r, 2, v); c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(name="Calibri", size=10, color=INK)
        r += 1
    ws.cell(r + 1, 1, "Headline finding").font = Font(name=SERIF, size=12, bold=True, color=NAVY)
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 6, end_column=2)
    hc = ws.cell(r + 2, 1, headline)
    hc.alignment = Alignment(wrap_text=True, vertical="top"); hc.font = Font(name="Calibri", size=10, color=INK)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95

    # ── 2) Assumptions & Decisions ───────────────────────────────────────────
    ws = wb.create_sheet("Assumptions & Decisions")
    dlog = TBL / "decisions_log.md"
    if dlog.exists():
        title(ws, "Assumptions & Decisions",
              "Every subjective judgment call + data note behind this deliverable", 1)
        paragraphs(ws, 4, dlog.read_text(encoding="utf-8"))
    else:
        title(ws, "Assumptions & Decisions",
              "Author Tables/decisions_log.md during the reasoning phase to populate this", 1)
        ws.cell(4, 1, "decisions_log.md not found — see references/threat-reasoning.md.")

    # ── 3) Methodology ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Methodology")
    title(ws, "Methodology", "How the analysis was produced, step by step", 1)
    pl = ", ".join(s.get("name", s["url"]) for s in cfg.get("parcel_sources", []))
    zl = ", ".join(s["jurisdiction"] for s in cfg.get("zoning_sources", []))
    method = f"""1. Subject resolution — the subject was geocoded and, where the mailing address mis-geocodes, overridden to a verified point ({sub.get('lat')}, {sub.get('lon')}). This anchors the map pin and all distances.
2. Analysis area — {area_desc}. Every parcel whose representative point falls inside this area is included.
3. Parcels — pulled from: {pl}. Normalized to a common schema (owner, situs, use code, value, year built).
4. Zoning — pulled from: {zl or '(none available)'}. Areas with no public zoning GIS are shown as a labeled 'No public zoning (data gap)' class — never faked as a district.
5. Land-use grouping — each appraisal use code is mapped to a land-use bucket (see Land-Use Grouping). Government / airport / school / church-owned land is routed to a Public/Institutional bucket by owner, so it does not read as developable vacant land.
6. Zoning grouping & MF threat — each zone code is grouped into a category and assigned a multifamily-supply threat tier (High = MF by-right, Medium = conditional, Low = not permitted, Unknown = planned-development). Parcels are joined to zoning by the containing polygon, falling back to ≥50% area overlap.
7. Vacant threat candidates — vacant parcels are mechanically filtered to genuinely developable land (drop roads/slivers by compactness, HOA/POA/condo common areas, churches/schools, gov/airport by owner, data-gap zoning, and parcels under {cfg.get('min_acres', 1.0)} ac), then adjacent commonly-owned parcels are clustered.
8. Reasoned Top-N — the final concerning-vacant ranking is a REASONED judgment over those candidates (not a mechanical score), with a written rationale per parcel. See Top Vacant Threats + Assumptions & Decisions.
9. Deliverables — this workbook + a self-contained interactive HTML viewer (Land Use / Zoning / Vacant Threats over satellite & street basemaps)."""
    paragraphs(ws, 4, method)

    # ── 4) Data Sources ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Data Sources")
    title(ws, "Data Sources", "Every public layer used, with its endpoint", 4)
    header(ws, 4, ["Type", "Jurisdiction / Name", "Endpoint", "Notes"])
    srows = []
    for s in cfg.get("parcel_sources", []):
        srows.append(["Parcels", s.get("name") or s.get("county", ""), s["url"],
                      f"confidence: {s.get('data_confidence', 'full')}"])
    for s in cfg.get("zoning_sources", []):
        srows.append(["Zoning", s["jurisdiction"], s["url"], s.get("note", "")])
    body(ws, 5, srows, [12, 28, 80, 30], wrapcols=(2, 3))

    # ── 5) Land-Use Grouping ─────────────────────────────────────────────────
    rows_lu = _csv(TBL / "land_use_buckets.csv")
    ws = wb.create_sheet("Land-Use Grouping")
    title(ws, "Land-Use Grouping",
          "How each appraisal use-code becomes a land-use category", 5)
    header(ws, 4, ["Land-Use Category", "Use Code", "What the code means", "Parcels in area", "Source"])
    lr = [[r["Bucket"], r["Code"], r["Description"], r.get("Parcels_in_area", ""), r["Source"]] for r in rows_lu]
    lc = [P.landuse_color(r["Bucket"]) for r in rows_lu]
    body(ws, 5, lr, [32, 14, 54, 14, 22], wrapcols=(2,), catcolorcol=0, catcolors=lc)

    # ── 6) Zoning Grouping ───────────────────────────────────────────────────
    rows_z = _csv(TBL / "zoning_categories.csv")
    zplain = cfg.get("zone_plain", {})
    ws = wb.create_sheet("Zoning Grouping")
    title(ws, "Zoning Grouping",
          "How each district is grouped and its apartment-supply threat", 6)
    header(ws, 4, ["Jurisdiction", "Zone Code", "What the district is", "Grouped Category",
                   "Apartment Threat", "What the threat means"])
    torder = {"High": 0, "Medium": 1, "Unknown": 2, "Low": 3}
    rows_z.sort(key=lambda r: (torder.get(r["MF Threat (if vacant)"], 9), r["Category"], r["Base Zone"]))
    zr = [[r["Jurisdiction"], r["Base Zone"], zplain.get(r["Base Zone"], r["Base Zone"]),
           r["Category"], r["MF Threat (if vacant)"],
           P.THREAT_DEF.get(threat_word(r["MF Threat (if vacant)"]), "")] for r in rows_z]
    body(ws, 5, zr, [22, 12, 28, 30, 14, 52], wrapcols=(2, 3, 5), threatcol=4)

    # ── 7) MF Threat Classification ──────────────────────────────────────────
    ws = wb.create_sheet("MF Threat Classification")
    title(ws, "Multifamily Supply-Threat Tiers",
          "What each threat tier means if the parcel were vacant or redeveloped", 2)
    header(ws, 4, ["Threat Tier", "Definition"])
    tr = [[t, P.THREAT_DEF[t]] for t in P.THREAT_ORDER]
    body(ws, 5, tr, [16, 96], wrapcols=(1,), threatcol=0)

    # ── 8) Parcels (the actual classified data) ──────────────────────────────
    pj = IN / "parcels_classified.geojson"
    if pj.exists():
        parcels = json.loads(pj.read_text())["features"]
        dev = set(json.loads((IN / "developable_accounts.json").read_text())) \
            if (IN / "developable_accounts.json").exists() else set()
        prows = []
        for f in parcels:
            p = f["properties"]
            prows.append([
                p.get("account"), p.get("owner"), p.get("situs"), p.get("county"),
                p.get("landuse_code"), p.get("bucket"), p.get("jurisdiction"),
                p.get("zone_plain"), p.get("zone_category"), p.get("mf_threat"),
                "Yes" if p.get("is_vacant") else "", "Yes" if p.get("account") in dev else "",
                p.get("acres"), p.get("dist_mi"), p.get("year_built")])
        prows.sort(key=lambda r: (r[13] if isinstance(r[13], (int, float)) else 9e9))
        lean_sheet("Parcels", f"All Parcels in Area ({len(prows):,})",
                   "Every parcel with its land use, zoning, and MF-supply threat — sorted by distance",
                   ["Account", "Owner", "Address", "County", "Use Code", "Land-Use Bucket",
                    "Jurisdiction", "Zoning", "Zone Category", "MF Threat", "Vacant?",
                    "Flagged Developable-Vacant", "Acres", "Dist (mi)", "Yr Built"],
                   prows,
                   [16, 28, 26, 12, 9, 24, 18, 20, 24, 12, 8, 14, 9, 9, 9], threatcol=9)

    # ── 9) Vacant — Developable (flagged threat set) ─────────────────────────
    vj = IN / "vacant_candidates.json"
    if vj.exists():
        cands = json.loads(vj.read_text())
        vrows = [[c.get("owner"), c.get("owner_type"), c.get("location"), c.get("zoning"),
                  c.get("zoning_threat_baseline"), c.get("acres"), c.get("parcel_count"),
                  c.get("min_dist_mi"), c.get("compactness"), c.get("lat"), c.get("lon"),
                  c.get("parcels")] for c in cands]
        vrows.sort(key=lambda r: (r[7] if isinstance(r[7], (int, float)) else 9e9))
        lean_sheet("Vacant - Developable", f"Developable Vacant Land Flagged ({len(vrows):,} sites)",
                   "Genuinely-developable vacant parcels after mechanical filters — the threat-map "
                   "set the reasoned Top-N is chosen from (zoning threat = base-zone baseline)",
                   ["Owner", "Owner Type", "Location", "Zoning (juris:code)", "Zone Threat (base)",
                    "Acres", "# Parcels", "Dist (mi)", "Compactness", "Lat", "Lon", "Parcel IDs"],
                   vrows,
                   [28, 20, 24, 26, 16, 9, 9, 9, 11, 11, 11, 30], threatcol=4)

    # ── 10) Top-N Vacant Threats ─────────────────────────────────────────────
    top = _csv(TBL / "top10_concerning_vacant.csv")
    ws = wb.create_sheet("Top Vacant Threats")
    title(ws, "Concerning Vacant Parcels (Reasoned)",
          "Developable land where new apartments could land — ranked by judgment, not a score", 9)
    header(ws, 4, ["#", "Owner", "Location", "Acres", "Dist (mi)", "Zoning",
                   "Apartment Threat", "Why it's concerning", "Parcels"])
    trows = [[r["rank"], r["OWNER1"], r["LOCATION"], r["acres"], r["dist_mi"],
              r["zoning_plain"], r["reasoned_threat"], r["rationale"], r["PARCEL"]] for r in top]
    end = body(ws, 5, trows, [4, 26, 22, 8, 9, 28, 14, 56, 22], wrapcols=(1, 5, 7, 8), threatcol=6)
    ws.merge_cells(start_row=end + 1, start_column=1, end_row=end + 1, end_column=9)
    ws.cell(end + 1, 1, "Ranking and threat are a reasoned judgment over mechanically-prepared "
            "candidates, not an automated score.").font = Font(name="Calibri", size=9, italic=True, color=GRAY)

    safe = "".join(ch if ch.isalnum() or ch in " -_" else "" for ch in name).strip()
    out = C.root(cfg) / f"{safe} - Land Use Analysis.xlsx"
    try:
        wb.save(out)
        print(f"wrote {out.name}  sheets={wb.sheetnames}")
    except PermissionError:
        alt = out.with_name(out.stem + " (UPDATED).xlsx")
        wb.save(alt)
        print(f"[original open] wrote {alt.name}")


if __name__ == "__main__":
    main()
